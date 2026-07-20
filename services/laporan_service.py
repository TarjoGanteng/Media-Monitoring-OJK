"""
services/laporan_service.py - Service untuk generate laporan PDF dan Excel

Fitur:
- Query data berita sesuai filter (tanggal, wilayah, topik, jenis_media)
- Generate chart server-side dengan matplotlib (base64 PNG)
- Render template HTML → PDF menggunakan WeasyPrint
- Generate Excel dengan openpyxl
- Manajemen riwayat laporan (CRUD)
"""

import os
import io
import base64
import logging
from datetime import datetime, date, timedelta
from collections import Counter
from sqlalchemy import func, desc, case

from database.extensions import db
from database.models import Berita, Laporan, User

logger = logging.getLogger(__name__)


# ==================== HELPER CHART ====================

def _render_chart_base64(fig) -> str:
    """Konversi matplotlib figure ke base64 PNG string untuk embed di HTML/PDF."""
    import matplotlib
    matplotlib.use("Agg")  # Non-interactive backend
    buf = io.BytesIO()
    fig.savefig(buf, format="png", bbox_inches="tight", dpi=120, facecolor="white")
    buf.seek(0)
    img_b64 = base64.b64encode(buf.read()).decode("utf-8")
    buf.close()
    import matplotlib.pyplot as plt
    plt.close(fig)
    return f"data:image/png;base64,{img_b64}"


def _chart_trend_harian(labels: list, total: list, positif: list, negatif: list, netral: list) -> str:
    """Chart line trend pemberitaan harian."""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import matplotlib.ticker as ticker

    fig, ax = plt.subplots(figsize=(8, 3.2))
    fig.patch.set_facecolor("white")
    ax.set_facecolor("#f8fafc")

    ax.plot(labels, total, color="#c8002d", linewidth=2.5, marker="o", markersize=5, label="Total", zorder=3)
    ax.fill_between(range(len(labels)), total, alpha=0.08, color="#c8002d")
    ax.plot(labels, positif, color="#16a34a", linewidth=1.8, marker="o", markersize=4, label="Positif", linestyle="--")
    ax.plot(labels, negatif, color="#dc2626", linewidth=1.8, marker="o", markersize=4, label="Negatif", linestyle="--")

    step = max(1, len(labels) // 6)
    ax.set_xticks(range(0, len(labels), step))
    ax.set_xticklabels([labels[i] for i in range(0, len(labels), step)], fontsize=8, rotation=15)
    ax.yaxis.set_major_locator(ticker.MaxNLocator(integer=True))
    ax.tick_params(axis="y", labelsize=8)
    ax.legend(fontsize=8, loc="upper left", framealpha=0.8)
    ax.grid(axis="y", linestyle="--", alpha=0.4)
    ax.spines[["top", "right"]].set_visible(False)
    ax.set_xlabel("Tanggal", fontsize=8)
    ax.set_ylabel("Jumlah Berita", fontsize=8)
    fig.tight_layout(pad=1.0)
    return _render_chart_base64(fig)


def _chart_donut_sentimen(positif: int, negatif: int, netral: int) -> str:
    """Chart donut untuk distribusi sentimen."""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    total = positif + negatif + netral
    if total == 0:
        positif, negatif, netral = 1, 1, 1

    sizes  = [positif, netral, negatif]
    colors = ["#16a34a", "#64748b", "#dc2626"]
    labels = ["Positif", "Netral", "Negatif"]

    fig, ax = plt.subplots(figsize=(4.5, 3.5))
    fig.patch.set_facecolor("white")
    wedges, texts, autotexts = ax.pie(
        sizes, colors=colors, autopct="%1.1f%%",
        startangle=90, wedgeprops=dict(width=0.55, edgecolor="white", linewidth=2),
        pctdistance=0.75, textprops={"fontsize": 9}
    )
    for at in autotexts:
        at.set_fontsize(8)
        at.set_color("white")
        at.set_fontweight("bold")

    ax.legend(wedges, labels, loc="lower center", fontsize=8,
              bbox_to_anchor=(0.5, -0.15), ncol=3, frameon=False)
    ax.set_title("", pad=0)
    fig.tight_layout()
    return _render_chart_base64(fig)


def _chart_bar_topik(topik_list: list) -> str:
    """Chart bar horizontal untuk top 5 topik."""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    if not topik_list:
        return ""

    labels = [t["topik"] for t in reversed(topik_list[:5])]
    values = [t["jumlah"] for t in reversed(topik_list[:5])]
    colors = ["#c8002d", "#e84a6d", "#f08099", "#f9c0cb", "#fde8ec"][:len(labels)]

    fig, ax = plt.subplots(figsize=(6, 3.2))
    fig.patch.set_facecolor("white")
    ax.set_facecolor("#f8fafc")
    bars = ax.barh(labels, values, color=colors, edgecolor="white", height=0.55)

    for bar, val in zip(bars, values):
        ax.text(val + max(values) * 0.01, bar.get_y() + bar.get_height() / 2,
                f"{val:,}", va="center", ha="left", fontsize=8, color="#374151")

    ax.set_xlabel("Jumlah Berita", fontsize=8)
    ax.tick_params(axis="y", labelsize=8)
    ax.tick_params(axis="x", labelsize=7)
    ax.spines[["top", "right", "left"]].set_visible(False)
    ax.grid(axis="x", linestyle="--", alpha=0.4)
    ax.set_xlim(0, max(values) * 1.18)
    fig.tight_layout(pad=1.0)
    return _render_chart_base64(fig)


def _chart_bar_kota(kota_list: list) -> str:
    """Chart bar horizontal untuk top 5 kota."""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    if not kota_list:
        return ""

    labels = [k["kota"] for k in reversed(kota_list[:5])]
    values = [k["jumlah"] for k in reversed(kota_list[:5])]
    colors = ["#1d4ed8", "#3b82f6", "#60a5fa", "#93c5fd", "#bfdbfe"][:len(labels)]

    fig, ax = plt.subplots(figsize=(6, 3.2))
    fig.patch.set_facecolor("white")
    ax.set_facecolor("#f8fafc")
    bars = ax.barh(labels, values, color=colors, edgecolor="white", height=0.55)

    for bar, val in zip(bars, values):
        ax.text(val + max(values) * 0.01, bar.get_y() + bar.get_height() / 2,
                f"{val:,}", va="center", ha="left", fontsize=8, color="#374151")

    ax.set_xlabel("Jumlah Berita", fontsize=8)
    ax.tick_params(axis="y", labelsize=8)
    ax.tick_params(axis="x", labelsize=7)
    ax.spines[["top", "right", "left"]].set_visible(False)
    ax.grid(axis="x", linestyle="--", alpha=0.4)
    ax.set_xlim(0, max(values) * 1.18)
    fig.tight_layout(pad=1.0)
    return _render_chart_base64(fig)


def _chart_trend_bulanan(labels: list, total: list, pct_positif: list) -> str:
    """Chart bar + line kombinasi untuk trend bulanan."""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import numpy as np

    fig, ax1 = plt.subplots(figsize=(7, 3.5))
    fig.patch.set_facecolor("white")
    ax1.set_facecolor("#f8fafc")

    x = np.arange(len(labels))
    bars = ax1.bar(x, total, color="#c8002d", alpha=0.75, width=0.55, label="Jumlah Berita")
    ax1.set_ylabel("Jumlah Berita", fontsize=8, color="#c8002d")
    ax1.tick_params(axis="y", labelcolor="#c8002d", labelsize=7)
    ax1.set_xticks(x)
    ax1.set_xticklabels(labels, fontsize=7.5, rotation=15)
    ax1.spines[["top", "right"]].set_visible(False)

    # Tambahkan label value di atas bar
    for bar, val in zip(bars, total):
        ax1.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + max(total) * 0.02,
                 f"{val:,}", ha="center", va="bottom", fontsize=7, color="#374151")

    # Secondary axis untuk % positif
    if pct_positif and any(p > 0 for p in pct_positif):
        ax2 = ax1.twinx()
        ax2.plot(x, pct_positif, color="#16a34a", linewidth=2, marker="o",
                 markersize=5, label="% Positif")
        ax2.set_ylabel("% Sentimen Positif", fontsize=8, color="#16a34a")
        ax2.tick_params(axis="y", labelcolor="#16a34a", labelsize=7)
        ax2.set_ylim(0, 100)
        ax2.spines[["top", "left"]].set_visible(False)
        for i, (xi, yi) in enumerate(zip(x, pct_positif)):
            ax2.text(xi, yi + 3, f"{yi:.0f}%", ha="center", fontsize=7, color="#16a34a")

    ax1.grid(axis="y", linestyle="--", alpha=0.3)
    fig.tight_layout(pad=1.0)
    return _render_chart_base64(fig)


def _chart_donut_kontribusi(jabar: int, nasional: int) -> str:
    """Chart donut kontribusi Jawa Barat vs Non-Jawa Barat."""
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    total = jabar + nasional
    if total == 0:
        jabar, nasional = 1, 1

    fig, ax = plt.subplots(figsize=(3.5, 3.5))
    fig.patch.set_facecolor("white")
    wedges, texts, autotexts = ax.pie(
        [jabar, nasional],
        colors=["#c8002d", "#e2e8f0"],
        autopct="%1.1f%%",
        startangle=90,
        wedgeprops=dict(width=0.55, edgecolor="white", linewidth=2),
        pctdistance=0.75,
    )
    for at in autotexts:
        at.set_fontsize(9)
        at.set_fontweight("bold")
        at.set_color("white")
    autotexts[1].set_color("#374151")

    ax.legend(wedges, ["Jawa Barat", "Nasional"], loc="lower center", fontsize=8,
              bbox_to_anchor=(0.5, -0.18), ncol=2, frameon=False)
    fig.tight_layout()
    return _render_chart_base64(fig)


# ==================== DATA QUERY ====================

def _date_str_expr(field):
    try:
        bind = db.session.get_bind()
        if bind and bind.dialect.name == "postgresql":
            return func.to_char(field, "YYYY-MM-DD")
    except Exception:
        pass
    return func.strftime("%Y-%m-%d", field)


def _year_str_expr(field):
    try:
        bind = db.session.get_bind()
        if bind and bind.dialect.name == "postgresql":
            return func.to_char(field, "YYYY")
    except Exception:
        pass
    return func.strftime("%Y", field)


class LaporanService:
    """Service untuk logika bisnis fitur laporan."""

    # ---- Nomor Laporan ----

    @staticmethod
    def generate_nomor_laporan() -> str:
        """Generate nomor laporan unik: LAP-MM-YYYY-NNN."""
        tahun = datetime.now().year
        yr_expr = _year_str_expr(Laporan.created_at)
        count = Laporan.query.filter(
            yr_expr == str(tahun)
        ).count() + 1
        return f"LAP-MM-{tahun}-{count:03d}"

    # ---- Data Laporan ----

    @staticmethod
    def get_data_laporan(
        tanggal_dari: date,
        tanggal_sampai: date,
        wilayah: str = None,
        topik: str = None,
        jenis_media: str = None,
    ) -> dict:
        """
        Mengumpulkan semua data yang diperlukan untuk laporan dari database.

        Returns:
            Dictionary berisi semua data: statistik, topik, kota, media, trend, berita, dll.
        """
        from datetime import datetime as _dt

        dt_dari = _dt.combine(tanggal_dari, _dt.min.time())
        dt_sampai = _dt.combine(tanggal_sampai, _dt.max.time())

        # Query base
        base_q = Berita.query.filter(
            Berita.status == "aktif",
            Berita.tanggal >= dt_dari,
            Berita.tanggal <= dt_sampai,
        )

        if wilayah and wilayah.lower() not in ["semua", "jawa barat", ""]:
            base_q = base_q.filter(Berita.wilayah.ilike(f"%{wilayah}%"))

        if topik and topik.lower() not in ["semua", ""]:
            base_q = base_q.filter(Berita.topik.ilike(f"%{topik}%"))

        if jenis_media and jenis_media.lower() == "lokal":
            base_q = base_q.filter(Berita.jenis_media == "Lokal")
        elif jenis_media and jenis_media.lower() == "non-lokal":
            base_q = base_q.filter(Berita.jenis_media == "Non-Lokal")

        # --- Statistik Utama ---
        total    = base_q.count()
        positif  = base_q.filter(Berita.sentimen == "Positif").count()
        negatif  = base_q.filter(Berita.sentimen == "Negatif").count()
        netral   = base_q.filter(Berita.sentimen == "Netral").count()
        pct_pos  = round(positif / total * 100, 1) if total > 0 else 0
        pct_neg  = round(negatif / total * 100, 1) if total > 0 else 0
        pct_net  = round(netral  / total * 100, 1) if total > 0 else 0

        statistik = {
            "total": total, "positif": positif, "negatif": negatif, "netral": netral,
            "pct_positif": pct_pos, "pct_negatif": pct_neg, "pct_netral": pct_net,
        }

        # --- Top 5 Topik ---
        topik_rows = (
            db.session.query(Berita.topik, func.count(Berita.id).label("jumlah"))
            .filter(Berita.status == "aktif", Berita.tanggal >= dt_dari,
                    Berita.tanggal <= dt_sampai, Berita.topik.isnot(None))
            .group_by(Berita.topik).order_by(desc("jumlah")).limit(5).all()
        )
        total_topik = sum(r.jumlah for r in topik_rows)
        topik_terbanyak = [
            {"topik": r.topik, "jumlah": r.jumlah,
             "pct": round(r.jumlah / total_topik * 100, 1) if total_topik else 0}
            for r in topik_rows
        ]

        # --- Top 5 Kota ---
        kota_rows = (
            db.session.query(Berita.wilayah, func.count(Berita.id).label("jumlah"))
            .filter(Berita.status == "aktif", Berita.tanggal >= dt_dari,
                    Berita.tanggal <= dt_sampai, Berita.wilayah.isnot(None))
            .group_by(Berita.wilayah).order_by(desc("jumlah")).limit(5).all()
        )
        kota_terbanyak = [{"kota": r.wilayah, "jumlah": r.jumlah} for r in kota_rows]
        kota_utama = kota_terbanyak[0]["kota"] if kota_terbanyak else "-"

        # --- Top 5 Media ---
        media_rows = (
            db.session.query(Berita.media, func.count(Berita.id).label("jumlah"))
            .filter(Berita.status == "aktif", Berita.tanggal >= dt_dari,
                    Berita.tanggal <= dt_sampai, Berita.media.isnot(None))
            .group_by(Berita.media).order_by(desc("jumlah")).limit(5).all()
        )
        media_teraktif = [{"media": r.media, "jumlah": r.jumlah} for r in media_rows]
        media_utama = media_teraktif[0]["media"] if media_teraktif else "-"

        # --- Topik paling banyak negatif ---
        topik_negatif_rows = (
            db.session.query(Berita.topik, func.count(Berita.id).label("jumlah"))
            .filter(Berita.status == "aktif", Berita.tanggal >= dt_dari,
                    Berita.tanggal <= dt_sampai, Berita.sentimen == "Negatif",
                    Berita.topik.isnot(None))
            .group_by(Berita.topik).order_by(desc("jumlah")).first()
        )
        topik_negatif_utama = topik_negatif_rows.topik if topik_negatif_rows else "-"
        topik_utama = topik_terbanyak[0]["topik"] if topik_terbanyak else "-"

        # --- Trend Harian ---
        hari = (tanggal_sampai - tanggal_dari).days + 1
        hari = min(hari, 30)  # Maks 30 hari untuk chart
        tgl_expr = _date_str_expr(Berita.tanggal)
        trend_rows = (
            db.session.query(
                tgl_expr.label("tgl"),
                func.count(Berita.id).label("total"),
                func.sum(case((Berita.sentimen == "Positif", 1), else_=0)).label("positif"),
                func.sum(case((Berita.sentimen == "Negatif", 1), else_=0)).label("negatif"),
                func.sum(case((Berita.sentimen == "Netral",  1), else_=0)).label("netral"),
            )
            .filter(Berita.status == "aktif", Berita.tanggal >= dt_dari, Berita.tanggal <= dt_sampai)
            .group_by(tgl_expr)
            .all()
        )
        trend_map = {r.tgl: r for r in trend_rows}
        date_range = [(tanggal_sampai - timedelta(days=i)) for i in range(hari - 1, -1, -1)]
        trend_labels, trend_total, trend_pos, trend_neg, trend_net = [], [], [], [], []
        for d in date_range:
            r = trend_map.get(d.strftime("%Y-%m-%d"))
            trend_labels.append(d.strftime("%d %b"))
            trend_total.append(int(r.total) if r else 0)
            trend_pos.append(int(r.positif) if r else 0)
            trend_neg.append(int(r.negatif) if r else 0)
            trend_net.append(int(r.netral) if r else 0)

        # --- Trend Bulanan ---
        from dateutil.relativedelta import relativedelta
        nama_bulan = ["","Jan","Feb","Mar","Apr","Mei","Jun","Jul","Agt","Sep","Okt","Nov","Des"]
        bulan_range = []
        curr = tanggal_dari.replace(day=1)
        while curr <= tanggal_sampai:
            bulan_range.append(curr)
            curr = (curr + relativedelta(months=1)).replace(day=1)

        trend_bulan_labels, trend_bulan_total, trend_bulan_pos_pct = [], [], []
        for b in bulan_range:
            bq = base_q.filter(Berita.bulan == b.month, Berita.tahun == b.year)
            t = bq.count()
            p = bq.filter(Berita.sentimen == "Positif").count()
            trend_bulan_labels.append(f"{nama_bulan[b.month]} {b.year}")
            trend_bulan_total.append(t)
            trend_bulan_pos_pct.append(round(p / t * 100, 1) if t > 0 else 0)

        # --- Kontribusi Jawa Barat vs Nasional ---
        total_nasional = Berita.query.filter(
            Berita.status == "aktif",
            Berita.tanggal >= dt_dari,
            Berita.tanggal <= dt_sampai,
        ).count()
        jabar_list = [
            "bandung", "bogor", "bekasi", "depok", "cirebon", "garut", "sukabumi",
            "tasikmalaya", "karawang", "cianjur", "cimahi", "sumedang", "majalengka",
            "subang", "purwakarta", "kuningan", "indramayu", "ciamis", "banjar",
            "pangandaran", "jawa barat"
        ]
        total_jabar_q = (
            db.session.query(func.count(Berita.id))
            .filter(Berita.status == "aktif", Berita.tanggal >= dt_dari,
                    Berita.tanggal <= dt_sampai,
                    func.lower(Berita.wilayah).in_(jabar_list))
            .scalar()
        )
        total_jabar = total_jabar_q or total  # fallback ke total filter
        kontribusi_pct = round(total_jabar / total_nasional * 100, 1) if total_nasional > 0 else 0
        total_non_jabar = max(0, total_nasional - total_jabar)

        # --- Berita Negatif Menonjol (top 5) ---
        berita_negatif = (
            Berita.query.filter(
                Berita.status == "aktif",
                Berita.tanggal >= dt_dari,
                Berita.tanggal <= dt_sampai,
                Berita.sentimen == "Negatif",
            )
            .order_by(desc(Berita.tanggal))
            .limit(5)
            .all()
        )

        # --- Daftar Berita Terbaru (10 berita untuk lampiran) ---
        berita_terbaru = (
            base_q.order_by(desc(Berita.tanggal)).limit(10).all()
        )

        # --- AI Ringkasan Eksekutif ---
        ringkasan_ai = LaporanService._generate_ringkasan_ai(
            statistik=statistik,
            topik_utama=topik_utama,
            topik_negatif=topik_negatif_utama,
            kota_utama=kota_utama,
            media_utama=media_utama,
            kontribusi_pct=kontribusi_pct,
        )

        return {
            "statistik": statistik,
            "topik_terbanyak": topik_terbanyak,
            "kota_terbanyak": kota_terbanyak,
            "media_teraktif": media_teraktif,
            "topik_utama": topik_utama,
            "topik_negatif_utama": topik_negatif_utama,
            "kota_utama": kota_utama,
            "media_utama": media_utama,
            "trend": {
                "labels": trend_labels, "total": trend_total,
                "positif": trend_pos, "negatif": trend_neg, "netral": trend_net,
            },
            "trend_bulanan": {
                "labels": trend_bulan_labels,
                "total": trend_bulan_total,
                "pct_positif": trend_bulan_pos_pct,
            },
            "kontribusi": {
                "jabar": total_jabar,
                "non_jabar": total_non_jabar,
                "nasional": total_nasional,
                "pct": kontribusi_pct,
            },
            "berita_negatif": berita_negatif,
            "berita_terbaru": berita_terbaru,
            "ringkasan_ai": ringkasan_ai,
        }

    @staticmethod
    def _generate_ringkasan_ai(
        statistik: dict,
        topik_utama: str,
        topik_negatif: str,
        kota_utama: str,
        media_utama: str,
        kontribusi_pct: float,
    ) -> str:
        """Generate ringkasan eksekutif via AIService. Fallback ke narasi statis."""
        try:
            from services.ai_service import gemini
            prompt = f"""Anda adalah analis media senior untuk OJK (Otoritas Jasa Keuangan) Jawa Barat, Indonesia.
Buat ringkasan eksekutif laporan pemberitaan dalam 3-4 kalimat bahasa Indonesia formal dan profesional.

Data laporan:
- Total berita: {statistik['total']}
- Sentimen: Positif {statistik['positif']} ({statistik['pct_positif']}%), Netral {statistik['netral']} ({statistik['pct_netral']}%), Negatif {statistik['negatif']} ({statistik['pct_negatif']}%)
- Topik terbanyak: {topik_utama}
- Topik negatif utama: {topik_negatif}
- Kota terbanyak: {kota_utama}
- Media teraktif: {media_utama}
- Kontribusi pemberitaan Jawa Barat: {kontribusi_pct}% dari total nasional

Tulis narasi 3-4 kalimat yang menyebutkan jumlah berita, sentimen dominan dan maknanya,
topik utama, dan kontribusi Jawa Barat. Jangan gunakan bullet, langsung paragraf."""

            res = gemini.generate_narasi(prompt, temperature=0.3)
            if res and len(res.strip()) > 50:
                return res.strip()
        except Exception as e:
            logger.warning(f"[LaporanService] AI ringkasan gagal: {e}")

        s = statistik
        sentimen_dom = (
            "positif" if s["positif"] >= s["negatif"] and s["positif"] >= s["netral"]
            else "negatif" if s["negatif"] >= s["positif"] else "netral"
        )
        return (
            f"Selama periode laporan ini, telah dimonitor sebanyak {s['total']:,} berita pemberitaan "
            f"mengenai OJK yang bersumber dari media online nasional dan lokal. "
            f"Pemberitaan didominasi oleh sentimen {sentimen_dom} sebesar {s[f'pct_{sentimen_dom}']:.1f}% "
            f"dengan topik terbanyak mengenai {topik_utama}. "
            f"Provinsi Jawa Barat menyumbang {kontribusi_pct:.1f}% dari total pemberitaan nasional "
            f"mengenai OJK pada periode ini."
        )

    @staticmethod
    def _generate_analisis_halaman_ai(section_name: str, data: dict, params: dict) -> str:
        """
        Generate analisis naratif profesional untuk tiap section laporan agar isi PDF penuh dan rapi.
        Menggunakan AIService (OpenAI/Cohere) dengan fallback teks profesional statis yang panjang.
        """
        from services.ai_service import gemini
        
        prompt = ""
        fallback = ""
        s = data["statistik"]
        periode = params.get("periode_label", "Periode ini")
        wilayah = params.get("wilayah", "Jawa Barat")

        if section_name == "trend_harian":
            labels = ", ".join(map(str, data["trend"]["labels"][:10]))
            total_trend = sum(data["trend"]["total"])
            prompt = f"""Anda adalah analis media senior OJK Jawa Barat. Berdasarkan data trend harian pemberitaan OJK Jabar berikut:
- Periode: {periode}
- Kota/Wilayah: {wilayah}
- Sampel Tanggal/Hari: {labels}
- Total volume dalam trend: {total_trend} berita

Tulis 2 paragraf narasi analisis formal, taktis, dan mendalam (minimal 100 kata) tanpa bullet point. 
Jelaskan pola fluktuasi pemberitaan harian ini, bagaimana sentimen publik merespons dinamika harian, 
dan apa implikasinya terhadap reputasi OJK Jawa Barat. Jangan sebutkan prompt ini."""
            
            fallback = (
                f"Analisis tren harian menunjukkan tingkat fluktuasi volume pemberitaan yang dinamis "
                f"di wilayah {wilayah} sepanjang periode {periode}. Grafik tren di atas merefleksikan reaksi media online "
                f"terhadap agenda publikasi resmi OJK serta isu-isu lokal yang muncul di masyarakat. "
                f"Lonjakan volume berita pada hari-hari tertentu umumnya dipicu oleh publikasi rilis pers resmi, "
                f"kegiatan edukasi keuangan tatap muka, atau adanya penindakan terhadap kasus keuangan ilegal.\n\n"
                f"Secara strategis, fluktuasi harian yang stabil menunjukkan bahwa humas OJK Jawa Barat berhasil "
                f"menjaga ritme komunikasi publik dengan pers lokal. Tingkat keterlibatan media yang konsisten ini "
                f"sangat penting untuk menjaga agar pesan edukasi dan regulasi OJK tetap tersampaikan ke masyarakat Jabar, "
                f"sekaligus memitigasi potensi penyebaran isu negatif secara cepat di media online."
            )

        elif section_name == "sentimen":
            prompt = f"""Anda adalah analis media senior OJK Jawa Barat. Analisislah distribusi sentimen pemberitaan OJK Jabar berikut:
- Total Berita: {s['total']}
- Positif: {s['positif']} ({s['pct_positif']}%)
- Netral: {s['netral']} ({s['pct_netral']}%)
- Negatif: {s['negatif']} ({s['pct_negatif']}%)

Tulis 2 paragraf analisis formal, objektif, dan profesional (minimal 100 kata) tanpa bullet point.
Jelaskan makna persentase sentimen ini bagi reputasi lembaga OJK, bagaimana dominasi sentimen mempengaruhi tingkat kepercayaan masyarakat,
dan rekomendasi untuk menanggulangi porsi berita bersentimen negatif agar tidak berkembang menjadi krisis kepercayaan."""
            
            fallback = (
                f"Distribusi sentimen pemberitaan OJK Provinsi Jawa Barat didominasi oleh sentimen positif sebesar {s['pct_positif']}% "
                f"dan netral sebesar {s['pct_netral']}%, sementara sentimen negatif mencatat porsi sebesar {s['pct_negatif']}%. "
                f"Dominasi sentimen positif dan netral mengindikasikan bahwa sebagian besar informasi yang dikonsumsi masyarakat "
                f"terkait OJK bersifat informatif, mendidik, dan mendukung kebijakan regulasi industri jasa keuangan. "
                f"Tingkat kepercayaan publik dapat dijaga dengan baik melalui publikasi yang konsisten.\n\n"
                f"Meskipun porsi sentimen negatif relatif terkendali pada {s['pct_negatif']}%, humas OJK Jawa Barat tetap perlu mewaspadai "
                f"isu-isu sensitif yang melatarbelakanginya. Penanganan yang proaktif melalui klarifikasi media, rilis pers bantahan yang cepat, "
                f"serta penyediaan kanal informasi satu pintu akan sangat membantu dalam meredam eskalasi sentimen negatif "
                f"agar tidak mengikis tingkat reputasi positif lembaga."
            )

        elif section_name == "topik":
            topik_str = ", ".join([f"{t['topik']} ({t['jumlah']} berita)" for t in data["topik_terbanyak"][:5]])
            prompt = f"""Anda adalah analis media senior OJK Jawa Barat. Analisislah 5 topik pemberitaan terpopuler berikut:
- Daftar Topik: {topik_str}
- Topik Utama: {data['topik_utama']}

Tulis 2 paragraf analisis formal dan strategis (minimal 100 kata) tanpa bullet point.
Jelaskan mengapa topik utama tersebut sangat mendominasi perhatian media online, bagaimana tingkat literasi keuangan masyarakat Jabar mempengaruhi sebaran topik ini,
dan langkah antisipatif apa yang harus diambil OJK untuk mengedukasi masyarakat terkait topik-topik krusial."""
            
            fallback = (
                f"Berdasarkan sebaran topik pemberitaan, isu mengenai '{data['topik_utama']}' menempati posisi teratas "
                f"dalam fokus pemberitaan media online di Jawa Barat. Tingginya frekuensi topik ini merefleksikan minat pembaca yang besar "
                f"serta urgensi penyebaran informasi terkait regulasi dan perlindungan konsumen. Media online merespons kebutuhan "
                f"informasi masyarakat Jawa Barat secara cepat melalui ulasan, opini, maupun pemberitaan langsung terkait isu tersebut.\n\n"
                f"OJK Jawa Barat perlu memanfaatkan momentum ketertarikan media pada topik-topik hangat ini dengan cara meluncurkan kampanye "
                f"edukasi yang lebih terstruktur. Penyediaan konten edukasi yang mudah dipahami, kolaborasi dengan akademisi serta komunitas lokal, "
                f"serta penyebaran konten kreatif di media digital akan mengoptimalkan pemahaman publik dan melindungi konsumen dari risiko keuangan."
            )

        elif section_name == "kota":
            kota_str = ", ".join([f"{k['kota']} ({k['jumlah']} berita)" for k in data["kota_terbanyak"][:5]])
            prompt = f"""Anda adalah analis media senior OJK Jawa Barat. Analisislah 5 daerah/kota pemberitaan OJK Jabar terbanyak berikut:
- Daftar Kota: {kota_str}
- Kota Utama: {data['kota_utama']}

Tulis 2 paragraf analisis formal (minimal 100 kata) tanpa bullet point.
Jelaskan dinamika sebaran geografis pemberitaan ini, mengapa daerah seperti {data['kota_utama']} menonjol secara volume pemberitaan,
dan bagaimana peta sebaran ini dapat membantu OJK Jabar dalam memfokuskan program sosialisasi keuangan daerah."""
            
            fallback = (
                f"Sebaran geografis pemberitaan menempatkan wilayah Kota/Kabupaten '{data['kota_utama']}' sebagai kontributor "
                f"pemberitaan terbanyak terkait OJK di Jawa Barat. Konsentrasi pemberitaan di kota-kota besar di Jawa Barat mencerminkan "
                f"tingginya aktivitas industri keuangan, kerapatan populasi, serta besarnya sebaran kantor media massa di wilayah tersebut. "
                f"Ini menunjukkan adanya hubungan erat antara aktivitas ekonomi daerah dengan volume publikasi informasi keuangan.\n\n"
                f"Informasi distribusi wilayah ini memberikan panduan strategis yang sangat berharga bagi OJK Jawa Barat dalam menyusun "
                f"prioritas program sosialisasi. Wilayah dengan volume berita rendah memerlukan pendekatan edukasi yang lebih aktif secara fisik, "
                f"sedangkan wilayah dengan volume tinggi seperti {data['kota_utama']} dapat dioptimalkan melalui publikasi media massa lokal "
                f"dan kerja sama strategis dengan pemerintah daerah setempat."
            )

        elif section_name == "media":
            media_str = ", ".join([f"{m['media']} ({m['jumlah']} berita)" for m in data["media_teraktif"][:5]])
            prompt = f"""Anda adalah analis media senior OJK Jawa Barat. Analisislah daftar media teraktif berikut:
- Daftar Media: {media_str}
- Media Utama: {data['media_utama']}

Tulis 2 paragraf analisis formal (minimal 100 kata) tanpa bullet point.
Jelaskan peran media teraktif tersebut dalam membentuk persepsi publik terhadap OJK Jawa Barat, pentingnya menjalin kemitraan media yang harmonis (media relations),
dan bagaimana Humas OJK Jabar dapat mendorong pemberitaan yang akurat, edukatif, dan berimbang melalui media pers pers tersebut."""
            
            fallback = (
                f"Aktivitas pemberitaan media didominasi oleh '{data['media_utama']}' yang secara konsisten menerbitkan berita "
                f"terkait OJK Provinsi Jawa Barat. Media-media teraktif ini memegang peran kunci sebagai agen penyebar informasi "
                f"dan pembentuk opini publik di tingkat daerah. Akurasi penulisan berita oleh jurnalis pada media-media tersebut "
                f"sangat menentukan tingkat pemahaman pembaca terhadap kebijakan OJK.\n\n"
                f"Untuk itu, divisi hubungan masyarakat OJK Jawa Barat direkomendasikan untuk terus membina kemitraan pers yang erat "
                f"dan berkelanjutan. Kegiatan rutin seperti media briefing, workshop jurnalisme keuangan, serta penyediaan siaran pers "
                f"yang komprehensif dan mudah diakses akan meminimalisir kesalahan penafsiran berita sekaligus mendorong terwujudnya "
                f"karya jurnalisme keuangan daerah yang edukatif dan berimbang."
            )

        elif section_name == "kontribusi":
            k = data["kontribusi"]
            prompt = f"""Anda adalah analis media senior OJK Jawa Barat. Analisislah kontribusi pemberitaan Jawa Barat dibanding Nasional berikut:
- Berita Jawa Barat: {k['jabar']} berita ({k['pct']}%)
- Berita Nasional OJK: {k['nasional']} berita

Tulis 2 paragraf analisis formal dan berbobot (minimal 100 kata) tanpa bullet point.
Jelaskan signifikansi kontribusi pemberitaan Jawa Barat sebesar {k['pct']}% dari total nasional, mengapa volume Jabar tergolong besar,
dan posisi strategis Kantor Regional/Provinsi Jawa Barat dalam mengawal komunikasi publik OJK pusat di daerah."""
            
            fallback = (
                f"Kontribusi volume pemberitaan dari wilayah Jawa Barat mencapai {k['pct']}% dari total nasional pemberitaan OJK. "
                f"Angka kontribusi yang signifikan ini mencerminkan posisi Jawa Barat sebagai salah satu pusat pertumbuhan ekonomi "
                f"dan industri jasa keuangan terbesar di Indonesia. Dinamika perkembangan ekonomi di Jawa Barat secara alami menarik "
                f"sorotan media nasional maupun daerah untuk meliput kinerja dan kebijakan OJK di provinsi ini.\n\n"
                f"Proporsi kontribusi sebesar {k['pct']}% ini juga menunjukkan tanggung jawab humas OJK Jawa Barat yang besar dalam menjaga "
                f"reputasi institusi secara nasional. Setiap riak isu komunikasi publik yang terjadi di Jawa Barat akan memberikan pengaruh langsung "
                f"pada persepsi publik terhadap OJK di tingkat pusat. Oleh karena itu, standardisasi pesan dan koordinasi komunikasi yang harmonis "
                f"antara OJK pusat dan Kantor Regional Jawa Barat menjadi faktor yang sangat krusial."
            )

        elif section_name == "trend_bulanan":
            bulanan_str = ", ".join([f"{data['trend_bulanan']['labels'][i]} ({data['trend_bulanan']['total'][i]} berita)" for i in range(len(data["trend_bulanan"]["labels"][:6]))])
            prompt = f"""Anda adalah analis media senior OJK Jawa Barat. Analisislah tren pemberitaan bulanan berikut:
- Data per Bulan: {bulanan_str}

Tulis 2 paragraf analisis formal (minimal 100 kata) tanpa bullet point.
Jelaskan tren kenaikan atau penurunan volume berita per bulan ini, faktor musiman atau rilis kebijakan apa yang mempengaruhinya,
dan proyeksi arah tren komunikasi publik OJK Jabar ke depan."""
            
            fallback = (
                f"Tren bulanan menunjukkan fluktuasi pemberitaan yang bervariasi sepanjang beberapa bulan terakhir. "
                f"Analisis grafik bulanan memperlihatkan pola volume berita yang dipengaruhi oleh agenda kerja OJK serta isu ekonomi musiman. "
                f"Kenaikan volume pada bulan tertentu umumnya sejalan dengan rilis laporan triwulanan perkembangan industri keuangan Jabar, "
                f"ataupun respon media terhadap rilis regulasi baru dari OJK Pusat.\n\n"
                f"Memahami tren bulanan ini membantu humas OJK Jawa Barat dalam merancang rencana komunikasi publik jangka menengah. "
                f"OJK Jabar dapat mengantisipasi bulan-bulan dengan volume pemberitaan rendah dengan menyusun program sosialisasi "
                f"mandiri, serta mempersiapkan respons yang matang untuk bulan-bulan sibuk demi mempertahankan sentimen pemberitaan "
                f"agar tetap bernilai positif bagi institusi."
            )

        elif section_name == "penutup":
            prompt = f"""Anda adalah analis media senior OJK Jawa Barat. Buatlah kesimpulan penutup laporan pemberitaan periode {periode} dengan total {s['total']} berita.
Tulis 2 paragraf penutup formal dan saran strategis jangka panjang bagi humas OJK Jabar (minimal 100 kata) tanpa bullet point."""
            
            fallback = (
                f"Sebagai kesimpulan, pemantauan media sepanjang periode {periode} menunjukkan bahwa kinerja komunikasi publik "
                f"OJK Jawa Barat berjalan dengan sangat baik dan efektif. Sebagian besar pemberitaan bernilai positif dan netral, "
                f"yang membuktikan bahwa informasi terkait regulasi dan edukasi keuangan tersampaikan ke publik Jawa Barat dengan akurat. "
                f"Kerja sama yang erat dengan media massa lokal menjadi fondasi utama keberhasilan reputasi positif ini.\n\n"
                f"Ke depan, humas OJK Jawa Barat disarankan untuk tetap konsisten memperluas jangkauan sosialisasi perlindungan konsumen, "
                f"terutama ke kota-kota satelit dan pedesaan di Jawa Barat. Penggunaan media sosial interaktif, kemitraan media yang solid, "
                f"serta respons cepat terhadap isu-isu negatif seperti investasi ilegal akan semakin memperkuat citra OJK sebagai lembaga kredibel "
                f"yang dipercayai sepenuhnya oleh masyarakat Jawa Barat."
            )

        if prompt:
            try:
                res = gemini.generate_narasi(prompt, temperature=0.5)
                if res and len(res.strip()) > 50:
                    return res.strip()
            except Exception as e:
                logger.warning(f"[LaporanService] Gagal generate narasi AI untuk {section_name}: {e}")
        
        return fallback


    # ---- Generate PDF ----

    @staticmethod
    def generate_pdf(data: dict, params: dict, output_dir: str) -> str:
        """
        Generate file PDF laporan menggunakan xhtml2pdf (pure Python, tanpa GTK+).
        Bekerja di semua platform: Windows, Linux, Mac, VPS.

        Args:
            data: Data laporan dari get_data_laporan()
            params: Parameter laporan (judul, periode, dll)
            output_dir: Direktori output

        Returns:
            Path file PDF yang dihasilkan
        """
        import io
        import base64
        from flask import render_template, current_app
        from xhtml2pdf import pisa

        # Helper: konversi file gambar statis ke base64 agar tidak ada
        # ketergantungan path eksternal (cross-platform safe)
        def img_to_b64(rel_static_path: str) -> str:
            try:
                abs_path = os.path.join(current_app.static_folder, rel_static_path)
                if not os.path.exists(abs_path):
                    return ""
                with open(abs_path, "rb") as fimg:
                    raw = fimg.read()
                ext = rel_static_path.rsplit(".", 1)[-1].lower()
                mime = "image/png" if ext == "png" else "image/jpeg"
                return f"data:{mime};base64,{base64.b64encode(raw).decode()}"
            except Exception as ex:
                logger.warning(f"[LaporanService] Gagal load gambar {rel_static_path}: {ex}")
                return ""

        # Siapkan chart base64
        charts = {}
        try:
            charts["trend_harian"] = _chart_trend_harian(
                data["trend"]["labels"],
                data["trend"]["total"],
                data["trend"]["positif"],
                data["trend"]["negatif"],
                data["trend"]["netral"],
            )
        except Exception as e:
            logger.warning(f"Chart trend harian gagal: {e}")

        try:
            s = data["statistik"]
            charts["donut_sentimen"] = _chart_donut_sentimen(
                s["positif"], s["negatif"], s["netral"]
            )
        except Exception as e:
            logger.warning(f"Chart donut sentimen gagal: {e}")

        try:
            charts["bar_topik"] = _chart_bar_topik(data["topik_terbanyak"])
        except Exception as e:
            logger.warning(f"Chart bar topik gagal: {e}")

        try:
            charts["bar_kota"] = _chart_bar_kota(data["kota_terbanyak"])
        except Exception as e:
            logger.warning(f"Chart bar kota gagal: {e}")

        try:
            charts["trend_bulanan"] = _chart_trend_bulanan(
                data["trend_bulanan"]["labels"],
                data["trend_bulanan"]["total"],
                data["trend_bulanan"]["pct_positif"],
            )
        except Exception as e:
            logger.warning(f"Chart trend bulanan gagal: {e}")

        try:
            k = data["kontribusi"]
            charts["donut_kontribusi"] = _chart_donut_kontribusi(k["jabar"], k["non_jabar"])
        except Exception as e:
            logger.warning(f"Chart donut kontribusi gagal: {e}")

        # Konversi gambar statis ke base64
        logo_b64   = img_to_b64("img/ojk-logo.png")
        kantor_b64 = img_to_b64("img/kantor-regional.png")

        # Generate narasi analisis AI panjang untuk tiap halaman agar PDF terisi penuh dan rapi
        analisis = {}
        analisis["trend_harian"] = LaporanService._generate_analisis_halaman_ai("trend_harian", data, params)
        analisis["sentimen"]     = LaporanService._generate_analisis_halaman_ai("sentimen", data, params)
        analisis["topik"]        = LaporanService._generate_analisis_halaman_ai("topik", data, params)
        analisis["kota"]         = LaporanService._generate_analisis_halaman_ai("kota", data, params)
        analisis["media"]        = LaporanService._generate_analisis_halaman_ai("media", data, params)
        analisis["kontribusi"]   = LaporanService._generate_analisis_halaman_ai("kontribusi", data, params)
        analisis["trend_bulanan"] = LaporanService._generate_analisis_halaman_ai("trend_bulanan", data, params)
        analisis["penutup"]      = LaporanService._generate_analisis_halaman_ai("penutup", data, params)

        # Render HTML template
        html_content = render_template(
            "laporan/laporan_pdf.html",
            data=data,
            params=params,
            charts=charts,
            logo_b64=logo_b64,
            kantor_b64=kantor_b64,
            analisis=analisis,
        )


        # xhtml2pdf: pure Python, no GTK, cross-platform
        pdf_buffer = io.BytesIO()
        pisa_status = pisa.CreatePDF(
            io.StringIO(html_content),
            dest=pdf_buffer,
            encoding="utf-8",
        )

        if pisa_status.err:
            raise RuntimeError(
                f"xhtml2pdf gagal membuat PDF (err={pisa_status.err}). "
                "Periksa log untuk detail."
            )

        # Simpan file
        os.makedirs(output_dir, exist_ok=True)
        nomor = params.get("nomor_laporan", "laporan").replace("/", "-")
        filename = f"{nomor}.pdf"
        filepath = os.path.join(output_dir, filename)
        with open(filepath, "wb") as f:
            f.write(pdf_buffer.getvalue())

        logger.info(f"[LaporanService] PDF berhasil dibuat (xhtml2pdf): {filepath}")
        return filepath


    # ---- Generate Excel ----

    @staticmethod
    def generate_excel(data: dict, params: dict, output_dir: str) -> str:
        """
        Generate file Excel laporan menggunakan openpyxl.

        Returns:
            Path file Excel yang dihasilkan
        """
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()

        # ---- Sheet 1: Ringkasan ----
        ws1 = wb.active
        ws1.title = "Ringkasan"

        RED      = "C8002D"
        DARK_RED = "8B0000"
        LIGHT_RED= "FFF0F0"
        GRAY     = "F1F5F9"
        DARK     = "1E293B"
        WHITE    = "FFFFFF"

        def hfill(hex_color):
            return PatternFill("solid", fgColor=hex_color)

        def hfont(bold=False, size=11, color="1E293B"):
            return Font(bold=bold, size=size, color=color, name="Calibri")

        def hborder():
            thin = Side(style="thin", color="CBD5E1")
            return Border(left=thin, right=thin, top=thin, bottom=thin)

        # Header judul
        ws1.merge_cells("A1:G1")
        ws1["A1"] = f"LAPORAN PEMBERITAAN OJK – {params.get('periode_label', '').upper()}"
        ws1["A1"].font = Font(bold=True, size=14, color=WHITE, name="Calibri")
        ws1["A1"].fill = hfill(RED)
        ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws1.row_dimensions[1].height = 36

        ws1.merge_cells("A2:G2")
        ws1["A2"] = f"Periode: {params.get('tanggal_dari_str', '')} s/d {params.get('tanggal_sampai_str', '')}  |  Wilayah: {params.get('wilayah', 'Jawa Barat')}  |  Nomor: {params.get('nomor_laporan', '')}"
        ws1["A2"].font = Font(size=9, color="64748B", name="Calibri")
        ws1["A2"].alignment = Alignment(horizontal="center")
        ws1.row_dimensions[2].height = 20

        # Statistik Utama
        row = 4
        ws1.merge_cells(f"A{row}:G{row}")
        ws1[f"A{row}"] = "STATISTIK UTAMA"
        ws1[f"A{row}"].font = hfont(True, 11, WHITE)
        ws1[f"A{row}"].fill = hfill(DARK_RED)
        ws1[f"A{row}"].alignment = Alignment(horizontal="left", indent=1)
        ws1.row_dimensions[row].height = 22

        stats_data = [
            ("Total Berita",     data["statistik"]["total"],     None),
            ("Berita Positif",   data["statistik"]["positif"],   f"{data['statistik']['pct_positif']}%"),
            ("Berita Netral",    data["statistik"]["netral"],     f"{data['statistik']['pct_netral']}%"),
            ("Berita Negatif",   data["statistik"]["negatif"],   f"{data['statistik']['pct_negatif']}%"),
            ("Kota Terbanyak",   data["kota_utama"],             None),
            ("Topik Terbanyak",  data["topik_utama"],            None),
            ("Media Teraktif",   data["media_utama"],            None),
        ]
        row += 1
        for i, (label, value, extra) in enumerate(stats_data):
            fill = hfill(LIGHT_RED if i % 2 == 0 else WHITE)
            for col in range(1, 8):
                ws1.cell(row=row, column=col).fill = fill
                ws1.cell(row=row, column=col).border = hborder()
            ws1.cell(row=row, column=1).value = label
            ws1.cell(row=row, column=1).font = hfont(True, 10, DARK)
            ws1.cell(row=row, column=3).value = value
            ws1.cell(row=row, column=3).font = hfont(False, 10)
            if extra:
                ws1.cell(row=row, column=4).value = extra
                ws1.cell(row=row, column=4).font = hfont(False, 10, "64748B")
            ws1.row_dimensions[row].height = 20
            row += 1

        # Top 5 Topik
        row += 1
        ws1.merge_cells(f"A{row}:G{row}")
        ws1[f"A{row}"] = "TOP 5 TOPIK PEMBERITAAN"
        ws1[f"A{row}"].font = hfont(True, 11, WHITE)
        ws1[f"A{row}"].fill = hfill(DARK_RED)
        ws1[f"A{row}"].alignment = Alignment(horizontal="left", indent=1)
        ws1.row_dimensions[row].height = 22
        row += 1

        ws1.cell(row=row, column=1).value = "No"
        ws1.cell(row=row, column=2).value = "Topik"
        ws1.cell(row=row, column=3).value = "Jumlah Berita"
        ws1.cell(row=row, column=4).value = "Persentase"
        for col in range(1, 5):
            ws1.cell(row=row, column=col).font = hfont(True, 10, WHITE)
            ws1.cell(row=row, column=col).fill = hfill(RED)
            ws1.cell(row=row, column=col).border = hborder()
            ws1.cell(row=row, column=col).alignment = Alignment(horizontal="center")
        ws1.row_dimensions[row].height = 20
        row += 1

        for i, t in enumerate(data["topik_terbanyak"], 1):
            fill = hfill(LIGHT_RED if i % 2 == 0 else WHITE)
            ws1.cell(row=row, column=1).value = i
            ws1.cell(row=row, column=2).value = t["topik"]
            ws1.cell(row=row, column=3).value = t["jumlah"]
            ws1.cell(row=row, column=4).value = f"{t['pct']}%"
            for col in range(1, 5):
                ws1.cell(row=row, column=col).fill = fill
                ws1.cell(row=row, column=col).border = hborder()
                ws1.cell(row=row, column=col).font = hfont(False, 10)
                ws1.cell(row=row, column=col).alignment = Alignment(horizontal="center" if col in [1, 3, 4] else "left")
            ws1.row_dimensions[row].height = 18
            row += 1

        # Column widths sheet 1
        for col_idx, width in [(1, 5), (2, 28), (3, 18), (4, 14), (5, 12), (6, 12), (7, 12)]:
            ws1.column_dimensions[get_column_letter(col_idx)].width = width

        # ---- Sheet 2: Daftar Berita ----
        ws2 = wb.create_sheet("Daftar Berita")

        headers = ["No", "Tanggal", "Judul", "Media", "Topik", "Sentimen", "Wilayah", "Link"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws2.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, size=10, color=WHITE, name="Calibri")
            cell.fill = hfill(RED)
            cell.border = hborder()
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws2.row_dimensions[1].height = 22

        sentimen_colors = {"Positif": "DCFCE7", "Negatif": "FEE2E2", "Netral": "F1F5F9"}
        for i, berita in enumerate(data["berita_terbaru"], 1):
            row_num = i + 1
            s_color = sentimen_colors.get(berita.sentimen, "FFFFFF")
            row_data = [
                i,
                berita.tanggal.strftime("%d/%m/%Y %H:%M") if berita.tanggal else "-",
                berita.judul or "-",
                berita.media or "-",
                berita.topik or "-",
                berita.sentimen or "-",
                berita.wilayah or "-",
                berita.link or "-",
            ]
            for col_idx, value in enumerate(row_data, 1):
                cell = ws2.cell(row=row_num, column=col_idx, value=value)
                cell.font = Font(size=9, name="Calibri")
                cell.border = hborder()
                cell.alignment = Alignment(vertical="top", wrap_text=(col_idx == 3))
                if col_idx == 6:
                    cell.fill = hfill(s_color)
                    cell.font = Font(bold=True, size=9, name="Calibri")
                elif i % 2 == 0:
                    cell.fill = hfill(GRAY)
            ws2.row_dimensions[row_num].height = 30

        # Column widths sheet 2
        for col_idx, width in [(1,4),(2,16),(3,55),(4,18),(5,20),(6,12),(7,18),(8,40)]:
            ws2.column_dimensions[get_column_letter(col_idx)].width = width

        # Simpan
        os.makedirs(output_dir, exist_ok=True)
        nomor = params.get("nomor_laporan", "laporan").replace("/", "-")
        filename = f"{nomor}.xlsx"
        filepath = os.path.join(output_dir, filename)
        wb.save(filepath)

        logger.info(f"[LaporanService] Excel berhasil dibuat: {filepath}")
        return filepath

    # ---- Riwayat Laporan ----

    @staticmethod
    def get_riwayat(limit: int = 20) -> list:
        """Ambil riwayat laporan terbaru."""
        return (
            Laporan.query.order_by(desc(Laporan.created_at))
            .limit(limit)
            .all()
        )

    @staticmethod
    def get_laporan_by_id(laporan_id: int):
        """Ambil satu laporan berdasarkan ID."""
        return db.session.get(Laporan, laporan_id)

    @staticmethod
    def hapus_laporan(laporan_id: int) -> tuple[bool, str]:
        """Hapus laporan beserta file PDF dan Excel-nya."""
        lap = db.session.get(Laporan, laporan_id)
        if not lap:
            return False, "Laporan tidak ditemukan."

        # Hapus file fisik
        for path in [lap.path_pdf, lap.path_excel]:
            if path and os.path.exists(path):
                try:
                    os.remove(path)
                except Exception as e:
                    logger.warning(f"Gagal hapus file {path}: {e}")

        try:
            db.session.delete(lap)
            db.session.commit()
            return True, f"Laporan {lap.nomor_laporan} berhasil dihapus."
        except Exception as e:
            db.session.rollback()
            return False, f"Gagal menghapus laporan: {e}"

    # ---- Helper untuk route (preview data sebelum generate) ----

    @staticmethod
    def get_preview_data(
        tanggal_dari: date,
        tanggal_sampai: date,
        wilayah: str = None,
        topik: str = None,
        jenis_media: str = None,
    ) -> dict:
        """
        Ambil data ringkasan cepat untuk preview sebelum generate laporan.
        Lebih ringan dari get_data_laporan() — tidak include AI/chart.
        """
        from datetime import datetime as _dt

        dt_dari   = _dt.combine(tanggal_dari, _dt.min.time())
        dt_sampai = _dt.combine(tanggal_sampai, _dt.max.time())

        base_q = Berita.query.filter(
            Berita.status == "aktif",
            Berita.tanggal >= dt_dari,
            Berita.tanggal <= dt_sampai,
        )
        if wilayah and wilayah.lower() not in ["semua", "jawa barat", ""]:
            base_q = base_q.filter(Berita.wilayah.ilike(f"%{wilayah}%"))
        if topik and topik.lower() not in ["semua", ""]:
            base_q = base_q.filter(Berita.topik.ilike(f"%{topik}%"))
        if jenis_media and jenis_media.lower() == "lokal":
            base_q = base_q.filter(Berita.jenis_media == "Lokal")
        elif jenis_media and jenis_media.lower() == "non-lokal":
            base_q = base_q.filter(Berita.jenis_media == "Non-Lokal")

        total   = base_q.count()
        positif = base_q.filter(Berita.sentimen == "Positif").count()
        negatif = base_q.filter(Berita.sentimen == "Negatif").count()
        netral  = base_q.filter(Berita.sentimen == "Netral").count()

        kota_row = (
            db.session.query(Berita.wilayah, func.count(Berita.id).label("j"))
            .filter(Berita.status == "aktif", Berita.tanggal >= dt_dari,
                    Berita.tanggal <= dt_sampai, Berita.wilayah.isnot(None))
            .group_by(Berita.wilayah).order_by(desc("j")).first()
        )
        topik_row = (
            db.session.query(Berita.topik, func.count(Berita.id).label("j"))
            .filter(Berita.status == "aktif", Berita.tanggal >= dt_dari,
                    Berita.tanggal <= dt_sampai, Berita.topik.isnot(None))
            .group_by(Berita.topik).order_by(desc("j")).first()
        )

        return {
            "total": total,
            "positif": positif,
            "negatif": negatif,
            "netral": netral,
            "pct_positif": round(positif / total * 100, 1) if total > 0 else 0,
            "pct_negatif": round(negatif / total * 100, 1) if total > 0 else 0,
            "pct_netral":  round(netral  / total * 100, 1) if total > 0 else 0,
            "kota_terbanyak": {"kota": kota_row.wilayah, "jumlah": kota_row.j} if kota_row else None,
            "topik_terbanyak": {"topik": topik_row.topik, "jumlah": topik_row.j} if topik_row else None,
        }
