"""
routes/laporan.py - Blueprint untuk fitur Laporan

Alur yang benar:
    1. POST /laporan/generate      → Generate ke folder TEMP, return temp_key + preview URL
                                     Belum masuk riwayat DB
    2. GET  /laporan/preview/<key> → Tampilkan PDF dari folder temp (inline)
    3. GET  /laporan/download/<key>/pdf|excel
                                   → Download file temp, SEKALIGUS simpan ke DB (masuk riwayat)
    4. GET  /laporan               → Halaman utama (menampilkan riwayat yang sudah di-download)
    5. GET  /laporan/view/<id>     → View PDF dari riwayat (sudah tersimpan)
    6. GET  /laporan/download/<id>/pdf|excel → Download dari riwayat
    7. POST /laporan/hapus/<id>    → Hapus laporan dari riwayat + file

Temp files: instance/laporan_temp/<uuid>/
Permanent:  instance/laporan/
"""

import os
import glob
import json
import shutil
import uuid
import logging
from datetime import datetime, date
from flask import (
    Blueprint, render_template, request, jsonify,
    send_file, abort, redirect, url_for, current_app
)
from flask_login import login_required, current_user

from database.extensions import db
from database.models import Laporan
from routes.auth import role_required
from services.laporan_service import LaporanService
from services.berita_service import BeritaService

logger = logging.getLogger(__name__)

bp = Blueprint("laporan", __name__)


# ==================== HELPER ====================

def _get_laporan_dir() -> str:
    """Direktori penyimpanan permanen laporan di instance/laporan/ (fallback ke /tmp jika read-only)."""
    import tempfile
    try:
        d = os.path.join(current_app.instance_path, "laporan")
        os.makedirs(d, exist_ok=True)
        return d
    except Exception:
        d = os.path.join(tempfile.gettempdir(), "laporan")
        os.makedirs(d, exist_ok=True)
        return d


def _get_temp_dir(temp_key: str = None) -> str:
    """Direktori temp untuk laporan yang belum di-download (fallback ke /tmp jika read-only)."""
    import tempfile
    try:
        base = os.path.join(current_app.instance_path, "laporan_temp")
        os.makedirs(base, exist_ok=True)
    except Exception:
        base = os.path.join(tempfile.gettempdir(), "laporan_temp")
        os.makedirs(base, exist_ok=True)

    if temp_key:
        d = os.path.join(base, temp_key)
        try:
            os.makedirs(d, exist_ok=True)
        except Exception:
            pass
        return d
    return base


def _validate_temp_key(temp_key: str) -> bool:
    """Validasi temp_key adalah UUID yang aman."""
    import re
    return bool(re.match(r'^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$', temp_key))


def _cleanup_old_temp(max_age_hours: int = 24):
    """Hapus folder temp yang lebih tua dari max_age_hours."""
    import time
    base = _get_temp_dir()
    if not os.path.exists(base):
        return
    now = time.time()
    for entry in os.scandir(base):
        if entry.is_dir():
            age = now - entry.stat().st_mtime
            if age > max_age_hours * 3600:
                try:
                    shutil.rmtree(entry.path)
                    logger.debug(f"[laporan] Temp dir dihapus: {entry.path}")
                except Exception:
                    pass


def _parse_tanggal(tgl_str: str):
    """Parse string YYYY-MM-DD menjadi date."""
    try:
        return datetime.strptime(tgl_str, "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return None


def _build_periode_label(jenis: str, tgl_dari: date, tgl_sampai: date, triwulan_val: str = None) -> str:
    """Buat label periode yang mudah dibaca."""
    BULAN = ["", "Januari", "Februari", "Maret", "April", "Mei", "Juni",
             "Juli", "Agustus", "September", "Oktober", "November", "Desember"]
    if jenis == "harian":
        return f"Harian ({tgl_dari.strftime('%d %b %Y')})"
    elif jenis == "mingguan":
        return f"Mingguan ({tgl_dari.strftime('%d %b')} – {tgl_sampai.strftime('%d %b %Y')})"
    elif jenis == "bulanan":
        return f"Bulanan ({BULAN[tgl_dari.month]} {tgl_dari.year})"
    elif jenis == "triwulan":
        q_map = {"1": ("I","Jan","Mar"), "2": ("II","Apr","Jun"),
                 "3": ("III","Jul","Sep"), "4": ("IV","Okt","Des")}
        q = q_map.get(str(triwulan_val), ("?","?","?"))
        return f"Triwulan {q[0]} ({q[1]}–{q[2]} {tgl_dari.year})"
    elif jenis == "tahunan":
        return f"Tahunan ({tgl_dari.year})"
    else:
        return f"Custom ({tgl_dari.strftime('%d %b %Y')} – {tgl_sampai.strftime('%d %b %Y')})"


def _load_temp_meta(temp_key: str) -> dict:
    """Load metadata JSON dari folder temp."""
    meta_path = os.path.join(_get_temp_dir(temp_key), "meta.json")
    if not os.path.exists(meta_path):
        return None
    with open(meta_path, "r", encoding="utf-8") as f:
        return json.load(f)


def _save_to_riwayat(temp_key: str) -> Laporan:
    """
    Pindahkan file dari temp ke permanent dan simpan ke DB.
    Jika sudah tersimpan (nomor sama), return record yang ada.
    """
    meta = _load_temp_meta(temp_key)
    if not meta:
        return None

    params = meta["params"]
    nomor = params["nomor_laporan"]

    # Cek apakah sudah pernah disimpan
    existing = Laporan.query.filter_by(nomor_laporan=nomor).first()
    if existing:
        return existing

    # Pindahkan file ke permanent dir
    temp_dir = _get_temp_dir(temp_key)
    perm_dir = _get_laporan_dir()

    def move_file(src_path):
        if not src_path or not os.path.exists(src_path):
            return None
        fn = os.path.basename(src_path)
        dest = os.path.join(perm_dir, fn)
        shutil.copy2(src_path, dest)
        return dest

    path_pdf   = move_file(meta.get("path_pdf"))
    path_excel = move_file(meta.get("path_excel"))

    # Parse tanggal
    tgl_dari   = datetime.strptime(params["tanggal_dari"], "%Y-%m-%d").date()
    tgl_sampai = datetime.strptime(params["tanggal_sampai"], "%Y-%m-%d").date()

    s = meta.get("statistik", {})

    laporan_record = Laporan(
        nomor_laporan  = nomor,
        judul          = params["judul"],
        periode_label  = params["periode_label"],
        jenis_periode  = params["jenis_periode"],
        tanggal_dari   = tgl_dari,
        tanggal_sampai = tgl_sampai,
        wilayah        = params.get("wilayah", "Jawa Barat"),
        topik          = params.get("topik"),
        jenis_media    = params.get("jenis_media", "semua"),
        path_pdf       = path_pdf,
        path_excel     = path_excel,
        dibuat_oleh    = meta.get("dibuat_oleh_id"),
        total_berita   = s.get("total", 0),
        total_positif  = s.get("positif", 0),
        total_negatif  = s.get("negatif", 0),
        total_netral   = s.get("netral", 0),
    )
    db.session.add(laporan_record)
    db.session.commit()

    logger.info(f"[laporan] Laporan {nomor} disimpan ke riwayat (id={laporan_record.id})")
    return laporan_record


# ==================== ROUTES ====================

@bp.route("/laporan")
@login_required
@role_required("super_admin", "pemimpin")
def index():
    """Halaman utama laporan: filter, preview, riwayat."""
    riwayat      = LaporanService.get_riwayat(limit=20)
    daftar_topik = BeritaService.get_daftar_topik()
    daftar_media = BeritaService.get_daftar_media()

    today        = date.today()
    default_dari = today.replace(day=1)

    return render_template(
        "laporan/index.html",
        riwayat       = riwayat,
        daftar_topik  = daftar_topik,
        daftar_media  = daftar_media,
        default_dari  = default_dari.strftime("%Y-%m-%d"),
        default_sampai= today.strftime("%Y-%m-%d"),
        active_page   = "laporan",
    )


@bp.route("/laporan/preview-data", methods=["POST"])
@login_required
@role_required("super_admin", "pemimpin")
def preview_data():
    """AJAX: ringkasan data sebelum generate."""
    try:
        body       = request.get_json(silent=True) or {}
        tgl_dari   = _parse_tanggal(body.get("tanggal_dari"))
        tgl_sampai = _parse_tanggal(body.get("tanggal_sampai"))

        if not tgl_dari or not tgl_sampai:
            return jsonify({"success": False, "error": "Tanggal tidak valid."}), 400
        if tgl_dari > tgl_sampai:
            return jsonify({"success": False, "error": "Tanggal dari harus sebelum tanggal sampai."}), 400

        preview = LaporanService.get_preview_data(
            tanggal_dari=tgl_dari, tanggal_sampai=tgl_sampai,
            wilayah=body.get("wilayah"), topik=body.get("topik"),
            jenis_media=body.get("jenis_media"),
        )
        return jsonify({"success": True, "data": preview})

    except Exception as e:
        logger.error(f"[laporan/preview-data] {e}", exc_info=True)
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/laporan/generate", methods=["POST"])
@login_required
@role_required("super_admin", "pemimpin")
def generate():
    """
    Generate laporan ke folder TEMP. BELUM masuk riwayat DB.
    Return: temp_key + URL preview + URL download.
    Laporan baru masuk riwayat saat user meng-klik Download.
    """
    try:
        # Bersihkan temp lama (>24 jam)
        _cleanup_old_temp()

        body          = request.get_json(silent=True) or {}
        jenis_periode = body.get("jenis_periode", "custom")
        tgl_dari      = _parse_tanggal(body.get("tanggal_dari"))
        tgl_sampai    = _parse_tanggal(body.get("tanggal_sampai"))
        wilayah       = body.get("wilayah", "Jawa Barat") or "Jawa Barat"
        topik         = body.get("topik") or None
        jenis_media   = body.get("jenis_media") or "semua"
        triwulan_val  = body.get("triwulan_val")

        if not tgl_dari or not tgl_sampai:
            return jsonify({"success": False, "error": "Tanggal tidak valid."}), 400

        # Build metadata
        nomor_laporan = LaporanService.generate_nomor_laporan()
        periode_label = _build_periode_label(jenis_periode, tgl_dari, tgl_sampai, triwulan_val)
        judul_laporan = f"Laporan Pemberitaan OJK – {periode_label}"

        params = {
            "nomor_laporan":      nomor_laporan,
            "judul":              judul_laporan,
            "periode_label":      periode_label,
            "jenis_periode":      jenis_periode,
            "tanggal_dari":       tgl_dari,
            "tanggal_sampai":     tgl_sampai,
            "tanggal_dari_str":   tgl_dari.strftime("%d %B %Y"),
            "tanggal_sampai_str": tgl_sampai.strftime("%d %B %Y"),
            "wilayah":            wilayah,
            "topik":              topik,
            "jenis_media":        jenis_media,
            "dibuat_oleh_nama":   current_user.nama_lengkap or current_user.username,
            "tanggal_cetak":      datetime.now().strftime("%d %B %Y, %H:%M WIB"),
        }

        # Kumpulkan data
        logger.info(f"[laporan/generate] Mulai generate {nomor_laporan}...")
        data = LaporanService.get_data_laporan(
            tanggal_dari=tgl_dari, tanggal_sampai=tgl_sampai,
            wilayah=wilayah if wilayah.lower() not in ["semua", "jawa barat"] else None,
            topik=topik, jenis_media=jenis_media,
        )

        # Output ke folder TEMP (bukan permanent)
        temp_key  = str(uuid.uuid4())
        temp_dir  = _get_temp_dir(temp_key)
        os.makedirs(temp_dir, exist_ok=True)

        logger.info(f"[laporan/generate] Generating PDF ke temp/{temp_key}...")
        path_pdf   = LaporanService.generate_pdf(data, params, temp_dir)

        logger.info(f"[laporan/generate] Generating Excel...")
        path_excel = LaporanService.generate_excel(data, params, temp_dir)

        # Simpan metadata ke JSON (untuk digunakan saat download)
        s = data["statistik"]
        meta = {
            "params": {
                **{k: (v.strftime("%Y-%m-%d") if hasattr(v, "strftime") else v)
                   for k, v in params.items()},
            },
            "statistik": {
                "total":   s["total"],
                "positif": s["positif"],
                "negatif": s["negatif"],
                "netral":  s["netral"],
            },
            "path_pdf":       path_pdf,
            "path_excel":     path_excel,
            "dibuat_oleh_id": current_user.id,
        }
        with open(os.path.join(temp_dir, "meta.json"), "w", encoding="utf-8") as mf:
            json.dump(meta, mf, ensure_ascii=False, indent=2)

        logger.info(f"[laporan/generate] Berhasil: {nomor_laporan} (temp_key={temp_key})")

        return jsonify({
            "success":            True,
            "temp_key":           temp_key,
            "nomor_laporan":      nomor_laporan,
            "preview_url":        url_for("laporan.preview_temp",       temp_key=temp_key),
            "excel_preview_url":  url_for("laporan.preview_excel_temp", temp_key=temp_key),
            "download_pdf_url":   url_for("laporan.download_temp",      temp_key=temp_key, fmt="pdf"),
            "download_excel_url": url_for("laporan.download_temp",      temp_key=temp_key, fmt="excel"),
            "statistik":          s,
        })

    except Exception as e:
        logger.error(f"[laporan/generate] Error: {e}", exc_info=True)
        return jsonify({"success": False, "error": str(e)}), 500


@bp.route("/laporan/preview-temp/<temp_key>")
@login_required
@role_required("super_admin", "pemimpin")
def preview_temp(temp_key: str):
    """
    Tampilkan PDF dari folder temp secara inline (untuk iframe preview).
    TIDAK menyimpan ke riwayat.
    """
    if not _validate_temp_key(temp_key):
        abort(400)

    temp_dir = _get_temp_dir(temp_key)
    pdfs = glob.glob(os.path.join(temp_dir, "*.pdf"))
    if not pdfs:
        abort(404)

    response = send_file(
        pdfs[0],
        mimetype="application/pdf",
        as_attachment=False,
    )
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["Content-Security-Policy"] = "default-src 'self'"
    return response


@bp.route("/laporan/download-temp/<temp_key>/<fmt>")
@login_required
@role_required("super_admin", "pemimpin")
def download_temp(temp_key: str, fmt: str):
    """
    Download file dari temp. Pada saat download inilah laporan BARU disimpan ke riwayat DB.
    fmt: 'pdf' atau 'excel'
    """
    if not _validate_temp_key(temp_key):
        abort(400)
    if fmt not in ("pdf", "excel"):
        abort(400)

    # Simpan ke riwayat saat pertama kali download
    try:
        laporan_record = _save_to_riwayat(temp_key)
    except Exception as e:
        logger.error(f"[laporan/download-temp] Gagal simpan riwayat: {e}", exc_info=True)
        laporan_record = None

    # Serve file
    if fmt == "pdf":
        # Coba ambil dari riwayat (permanent), fallback ke temp
        if laporan_record and laporan_record.path_pdf and os.path.exists(laporan_record.path_pdf):
            path = laporan_record.path_pdf
        else:
            temp_dir = _get_temp_dir(temp_key)
            pdfs = glob.glob(os.path.join(temp_dir, "*.pdf"))
            if not pdfs:
                abort(404)
            path = pdfs[0]

        nomor = laporan_record.nomor_laporan if laporan_record else "Laporan"
        safe_name = nomor.replace("/", "-") + ".pdf"
        response = send_file(
            path,
            mimetype="application/pdf",
            as_attachment=True,
            download_name=safe_name,
        )
        response.headers["X-Content-Type-Options"] = "nosniff"
        response.headers["Content-Disposition"] = f'attachment; filename="{safe_name}"'
        return response

    else:  # excel
        if laporan_record and laporan_record.path_excel and os.path.exists(laporan_record.path_excel):
            path = laporan_record.path_excel
        else:
            temp_dir = _get_temp_dir(temp_key)
            xlsxs = glob.glob(os.path.join(temp_dir, "*.xlsx"))
            if not xlsxs:
                abort(404)
            path = xlsxs[0]

        nomor = laporan_record.nomor_laporan if laporan_record else "Laporan"
        safe_name = nomor.replace("/", "-") + ".xlsx"
        response = send_file(
            path,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=safe_name,
        )
        response.headers["X-Content-Type-Options"] = "nosniff"
        response.headers["Content-Disposition"] = f'attachment; filename="{safe_name}"'
        return response


# ==================== EXCEL PREVIEW HELPER ====================

def _excel_to_html(xlsx_path: str, nomor: str = "") -> str:
    """Konversi file Excel ke HTML table untuk preview di iframe."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception as e:
        return f"<p style='color:red;padding:2rem;'>Gagal membaca Excel: {e}</p>"

    sheets_html = []
    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        header = rows[0]
        data   = rows[1:]

        thead = "<thead><tr>" + "".join(
            f'<th>{(str(c) if c is not None else "")}' for c in header
        ) + "</tr></thead>"

        tbody_rows = []
        for i, row in enumerate(data):
            cls = 'class="even"' if i % 2 == 1 else ''
            cells = "".join(
                f'<td>{(str(c) if c is not None else "")}' for c in row
            )
            tbody_rows.append(f"<tr {cls}>{cells}</tr>")
        tbody = "<tbody>" + "".join(tbody_rows) + "</tbody>"

        sheets_html.append(f"""
          <div class="sheet-section">
            <div class="sheet-title">{sheetname}</div>
            <div class="table-wrap">
              <table>{thead}{tbody}</table>
            </div>
          </div>
        """)

    wb.close()

    newline = "\n"
    return f"""<!DOCTYPE html>
<html lang="id">
<head>
<meta charset="UTF-8">
<title>Preview Excel — {nomor}</title>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{
    font-family: -apple-system, 'Segoe UI', sans-serif;
    font-size: 11px;
    background: #f8fafc;
    color: #1e293b;
    padding: 0;
  }}
  .top-bar {{
    background: #16a34a;
    color: white;
    padding: 8px 16px;
    font-size: 12px;
    font-weight: 700;
    display: flex;
    align-items: center;
    gap: 8px;
    position: sticky;
    top: 0;
    z-index: 10;
  }}
  .top-bar .badge {{
    background: rgba(255,255,255,0.2);
    padding: 2px 10px;
    border-radius: 12px;
    font-size: 10px;
  }}
  .sheet-section {{
    margin: 12px;
  }}
  .sheet-title {{
    font-size: 11px;
    font-weight: 700;
    color: #16a34a;
    text-transform: uppercase;
    letter-spacing: 0.05em;
    margin-bottom: 6px;
    padding-bottom: 4px;
    border-bottom: 2px solid #16a34a;
  }}
  .table-wrap {{
    overflow-x: auto;
    border: 1px solid #e2e8f0;
    border-radius: 6px;
    background: white;
  }}
  table {{
    width: 100%;
    border-collapse: collapse;
    font-size: 10.5px;
  }}
  thead tr {{
    background: #f0fdf4;
  }}
  th {{
    padding: 6px 10px;
    text-align: left;
    font-weight: 700;
    font-size: 10px;
    color: #15803d;
    text-transform: uppercase;
    border-bottom: 2px solid #bbf7d0;
    white-space: nowrap;
  }}
  td {{
    padding: 5px 10px;
    border-bottom: 1px solid #f1f5f9;
    white-space: nowrap;
    max-width: 200px;
    overflow: hidden;
    text-overflow: ellipsis;
  }}
  tr.even td {{ background: #f8fffe; }}
  tr:hover td {{ background: #f0fdf4; }}
</style>
</head>
<body>
  <div class="top-bar">
    <svg width="16" height="16" viewBox="0 0 24 24" fill="white">
      <path d="M14 2H6a2 2 0 00-2 2v16a2 2 0 002 2h12a2 2 0 002-2V8l-6-6z"/>
      <path d="M14 2v6h6" fill="none" stroke="white" stroke-width="2"/>
    </svg>
    Preview Excel
    <span class="badge">{nomor}</span>
  </div>
  {newline.join(sheets_html)}
</body>
</html>"""


@bp.route("/laporan/preview-excel-temp/<temp_key>")
@login_required
@role_required("super_admin", "pemimpin")
def preview_excel_temp(temp_key: str):
    """Preview Excel dari folder temp sebagai HTML table."""
    if not _validate_temp_key(temp_key):
        abort(400)
    temp_dir = _get_temp_dir(temp_key)
    xlsxs = glob.glob(os.path.join(temp_dir, "*.xlsx"))
    if not xlsxs:
        abort(404)
    html = _excel_to_html(xlsxs[0], "Preview")
    from flask import Response
    return Response(html, mimetype="text/html")


@bp.route("/laporan/preview-excel/<int:laporan_id>")
@login_required
@role_required("super_admin", "pemimpin")
def preview_excel_riwayat(laporan_id: int):
    """Preview Excel dari riwayat sebagai HTML table."""
    lap = LaporanService.get_laporan_by_id(laporan_id)
    if not lap or not lap.path_excel or not os.path.exists(lap.path_excel):
        abort(404)
    html = _excel_to_html(lap.path_excel, lap.nomor_laporan)
    from flask import Response
    return Response(html, mimetype="text/html")


# ==================== RIWAYAT ROUTES ====================

@bp.route("/laporan/view/<int:laporan_id>")
@login_required
@role_required("super_admin", "pemimpin")
def view_pdf(laporan_id: int):
    """Tampilkan PDF dari riwayat secara inline."""
    lap = LaporanService.get_laporan_by_id(laporan_id)
    if not lap or not lap.path_pdf or not os.path.exists(lap.path_pdf):
        abort(404)
    response = send_file(lap.path_pdf, mimetype="application/pdf",
                         as_attachment=False,
                         download_name=f"{lap.nomor_laporan}.pdf")
    response.headers["X-Content-Type-Options"] = "nosniff"
    return response


@bp.route("/laporan/download/<int:laporan_id>/pdf")
@login_required
@role_required("super_admin", "pemimpin")
def download_pdf(laporan_id: int):
    """Download PDF dari riwayat."""
    lap = LaporanService.get_laporan_by_id(laporan_id)
    if not lap or not lap.path_pdf or not os.path.exists(lap.path_pdf):
        abort(404)
    safe_name = lap.nomor_laporan.replace("/", "-") + ".pdf"
    response = send_file(lap.path_pdf, mimetype="application/pdf",
                         as_attachment=True,
                         download_name=safe_name)
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["Content-Disposition"] = f'attachment; filename="{safe_name}"'
    return response


@bp.route("/laporan/download/<int:laporan_id>/excel")
@login_required
@role_required("super_admin", "pemimpin")
def download_excel(laporan_id: int):
    """Download Excel dari riwayat."""
    lap = LaporanService.get_laporan_by_id(laporan_id)
    if not lap or not lap.path_excel or not os.path.exists(lap.path_excel):
        abort(404)
    safe_name = lap.nomor_laporan.replace("/", "-") + ".xlsx"
    response = send_file(
        lap.path_excel,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=safe_name,
    )
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["Content-Disposition"] = f'attachment; filename="{safe_name}"'
    return response


@bp.route("/laporan/hapus/<int:laporan_id>", methods=["POST"])
@login_required
@role_required("super_admin", "pemimpin")
def hapus(laporan_id: int):
    """Hapus laporan dari riwayat + file fisik."""
    ok, msg = LaporanService.hapus_laporan(laporan_id)
    if request.is_json:
        return jsonify({"success": ok, "message": msg})
    return redirect(url_for("laporan.index"))


@bp.route("/laporan/riwayat-json")
@login_required
@role_required("super_admin", "pemimpin")
def riwayat_json():
    """AJAX: ambil daftar riwayat terbaru (untuk update tabel tanpa reload halaman)."""
    riwayat = LaporanService.get_riwayat(limit=20)
    data = []
    for r in riwayat:
        has_pdf   = bool(r.path_pdf   and os.path.exists(r.path_pdf))
        has_excel = bool(r.path_excel and os.path.exists(r.path_excel))
        data.append({
            "id":                r.id,
            "nomor_laporan":     r.nomor_laporan,
            "judul":             r.judul,
            "periode_label":     r.periode_label,
            "wilayah":           r.wilayah or "Jawa Barat",
            "total_berita":      r.total_berita,
            "created_at":        r.created_at.strftime("%d %b %Y %H:%M") if r.created_at else "-",
            "view_url":          url_for("laporan.view_pdf",          laporan_id=r.id) if has_pdf   else None,
            "download_pdf":      url_for("laporan.download_pdf",      laporan_id=r.id) if has_pdf   else None,
            "download_excel":    url_for("laporan.download_excel",    laporan_id=r.id) if has_excel else None,
            "excel_preview_url": url_for("laporan.preview_excel_riwayat", laporan_id=r.id) if has_excel else None,
            "hapus_url":         url_for("laporan.hapus",             laporan_id=r.id),
        })
    return jsonify({"success": True, "riwayat": data})

